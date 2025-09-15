'''
upgraded backtester with per-strategy Risk of Ruin analysis.

Now you can supply a dict of thresholds (in R units) per strategy. If a strategy is missing from the dict, it falls back to the global threshold.

Summary table export
After each run you’ll get a CSV (or Excel) with all stats per strategy + portfolio:

Number of trades
Win rate
Expectancy (R)
Max drawdown (R)
Risk of ruin %

Summary Table Export

Summary table with per-strategy + portfolio metrics.
Auto-export to strategy_summary.csv.
Includes RuinProb% from Monte Carlo simulation.

Excel reports with multiple sheets:

Summary → all per-strategy + portfolio stats.
Trades → every trade reconstructed (entry, exit, R).
Equity → portfolio equity curve (and optionally per-strategy curves).

Excel + PNG Export

Timestamped Reports

Excel + PNG + MonteCarlo Export

Configurable Monte Carlo Runs

Interactive Mode (no cli)

Equity + MC Histogram PNG Export

Excel with Embedded Charts + PNGs

Toggle Chart Embedding

Config File Support

Auto Generate Config

Excel Report Includes Config Sheet

Config + Metadata in Excel

Config sheet:
  Date/time of run
  Monte Carlo runs
  Global ruin threshold
  Embed toggle
  Trades.csv row count
  Strategies detected (comma-separated)
  Per-strategy ruin thresholds

'''

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from itertools import groupby
from datetime import datetime
import argparse
import io
import json, yaml
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

LEDGER_FILE = "trades.csv"
CONFIG_FILE = "config.yaml"

DEFAULT_CONFIG = {
    "montecarlo_runs": 1000,
    "ruin_threshold": -10,
    "embed_charts": True,
    "ruin_thresholds": {
        "180PC": -6,
        "T-Wave": -12,
        "Breakfast": -8
    }
}

def ensure_config():
    """Ensure config.yaml exists, else create with defaults."""
    cfg_path = Path(CONFIG_FILE)
    if not cfg_path.exists():
        with open(cfg_path, "w") as f:
            yaml.safe_dump(DEFAULT_CONFIG, f)
        print(f"⚙️  Default config created → {CONFIG_FILE}")
        return DEFAULT_CONFIG
    else:
        with open(cfg_path, "r") as f:
            if cfg_path.suffix == ".json":
                return json.load(f)
            else:
                return yaml.safe_load(f)

def load_trades(file=LEDGER_FILE):
    return pd.read_csv(file, parse_dates=["timestamp"])

def build_trade_pairs(df: pd.DataFrame):
    trades, open_trades = [], {}
    for _, row in df.iterrows():
        strat = row["strategy"]
        if row["event"] == "entry":
            open_trades[strat] = {
                "timestamp": row["timestamp"],
                "strategy": strat,
                "dir": row["dir"],
                "entry": row["entry"],
                "sl": row["sl"],
                "tp": row["tp"],
            }
        elif row["event"] in ["sl", "tp", "trail"]:
            if strat in open_trades:
                tr = open_trades.pop(strat)
                tr["exit_event"] = row["event"]
                tr["exit_price"] = row["price"]
                trades.append(tr)
    return pd.DataFrame(trades)

def compute_drawdown(equity):
    peaks = np.maximum.accumulate(equity)
    dd = equity - peaks
    max_dd = dd.min() if len(dd) else 0
    durations, cur_dur = [], 0
    for d in dd:
        if d < 0: cur_dur += 1
        elif cur_dur > 0:
            durations.append(cur_dur)
            cur_dur = 0
    avg_dd_dur = np.mean(durations) if durations else 0
    return max_dd, avg_dd_dur

def save_chart_as_image(fig, filename):
    fig.savefig(filename, dpi=300, bbox_inches="tight")
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    return buf

def analyze_trades(trades: pd.DataFrame, montecarlo_runs=1000,
                   ruin_threshold=-10, ruin_thresholds=None,
                   embed_charts=True):
    trades = trades.copy()
    trades["risk"] = (trades["entry"] - trades["sl"]).abs()
    trades["pnl"] = trades.apply(
        lambda r: (r["exit_price"] - r["entry"]) if r["dir"] == "long"
        else (r["entry"] - r["exit_price"]), axis=1)
    trades["R"] = trades["pnl"] / trades["risk"]

    summary = []
    mc_summary = []
    mc_results = pd.DataFrame()

    # === Portfolio stats ===
    trades = trades.sort_values("timestamp")
    trades["portfolio_equity"] = trades["R"].cumsum()
    max_dd, avg_dd_dur = compute_drawdown(trades["portfolio_equity"].values)
    winrate = (trades["R"] > 0).mean()
    expectancy = trades["R"].mean()
    summary.append({
        "Strategy": "Portfolio",
        "Trades": len(trades),
        "WinRate%": winrate*100,
        "ExpectancyR": expectancy,
        "MaxDD(R)": max_dd,
        "AvgDD_Dur": avg_dd_dur,
        "RuinProb%": None,
    })

    # === Strategy stats ===
    grouped = trades.groupby("strategy")
    equity_curves = {}
    for strat, g in grouped:
        g = g.sort_values("timestamp")
        g["equity"] = g["R"].cumsum()
        equity_curves[strat] = g[["timestamp", "equity"]].copy()
        max_dd, avg_dd_dur = compute_drawdown(g["equity"].values)
        winrate = (g["R"] > 0).mean()
        expectancy = g["R"].mean()
        summary.append({
            "Strategy": strat,
            "Trades": len(g),
            "WinRate%": winrate*100,
            "ExpectancyR": expectancy,
            "MaxDD(R)": max_dd,
            "AvgDD_Dur": avg_dd_dur,
            "RuinProb%": None,
        })

    # === Monte Carlo (Portfolio + per-strategy ruin) ===
    if montecarlo_runs > 0:
        results = []
        ruin_count = 0
        for _ in range(montecarlo_runs):
            shuffled = trades["R"].sample(frac=1, replace=False).reset_index(drop=True)
            eq = shuffled.cumsum()
            results.append(eq.iloc[-1])
            if eq.min() <= ruin_threshold:
                ruin_count += 1
        ruin_prob = ruin_count / montecarlo_runs
        for s in summary:
            if s["Strategy"] == "Portfolio":
                s["RuinProb%"] = ruin_prob*100
        mc_summary.append({
            "Strategy": "Portfolio",
            "RuinThreshold": ruin_threshold,
            "RuinProb%": ruin_prob*100,
            "MedianFinalR": np.median(results),
            "P05": np.percentile(results, 5),
            "P95": np.percentile(results, 95)
        })
        mc_results["Portfolio_FinalEquity"] = results

        if ruin_thresholds:
            for strat, g in grouped:
                if len(g) == 0: continue
                strat_thr = ruin_thresholds.get(strat, ruin_threshold)
                results = []
                ruin_count = 0
                for _ in range(montecarlo_runs):
                    shuffled = g["R"].sample(frac=1, replace=False).reset_index(drop=True)
                    eq = shuffled.cumsum()
                    results.append(eq.iloc[-1])
                    if eq.min() <= strat_thr:
                        ruin_count += 1
                ruin_prob = ruin_count / montecarlo_runs
                for s in summary:
                    if s["Strategy"] == strat:
                        s["RuinProb%"] = ruin_prob*100
                mc_summary.append({
                    "Strategy": strat,
                    "RuinThreshold": strat_thr,
                    "RuinProb%": ruin_prob*100,
                    "MedianFinalR": np.median(results),
                    "P05": np.percentile(results, 5),
                    "P95": np.percentile(results, 95)
                })
                mc_results[strat+"_FinalEquity"] = results

    # === Timestamped filenames ===
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    report_file = f"strategy_report_{ts}.xlsx"
    curve_file  = f"equity_curve_{ts}.png"
    mc_file     = f"mc_histogram_{ts}.png"

    # === Save Excel report ===
    df_summary = pd.DataFrame(summary)
    df_mc_summary = pd.DataFrame(mc_summary)

    # Runtime config/metadata
    runtime_info = [
        {"Parameter": "Run Timestamp", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Parameter": "Monte Carlo Runs", "Value": montecarlo_runs},
        {"Parameter": "Global Ruin Threshold", "Value": ruin_threshold},
        {"Parameter": "Embed Charts", "Value": embed_charts},
        {"Parameter": "Trades.csv Rows", "Value": len(trades)},
        {"Parameter": "Strategies Detected", "Value": ", ".join(sorted(trades["strategy"].unique()))},
    ]
    if ruin_thresholds:
        for k,v in ruin_thresholds.items():
            runtime_info.append({"Parameter": f"RuinThreshold_{k}", "Value": v})
    df_config = pd.DataFrame(runtime_info)

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        trades.to_excel(writer, sheet_name="Trades", index=False)
        eq_df = trades[["timestamp", "portfolio_equity"]].rename(columns={"portfolio_equity":"Portfolio"})
        for strat, eq in equity_curves.items():
            eq_df = eq_df.merge(eq, on="timestamp", how="outer", suffixes=("","_"+strat))
        eq_df.to_excel(writer, sheet_name="Equity", index=False)
        if not df_mc_summary.empty:
            df_mc_summary.to_excel(writer, sheet_name="MonteCarlo", index=False)
            mc_results.to_excel(writer, sheet_name="MC_Distribution", index=False)
        df_config.to_excel(writer, sheet_name="Config", index=False)

    # === Charts ===
    fig1 = plt.figure(figsize=(12,7))
    for strat, g in grouped:
        plt.plot(g["timestamp"], g["equity"], marker="o", label=strat)
    plt.plot(trades["timestamp"], trades["portfolio_equity"],
             color="black", linewidth=2.5, label="Portfolio Total")
    plt.title("Equity Curves by Strategy + Portfolio (in R units)")
    plt.legend()
    plt.grid(True)
    buf1 = save_chart_as_image(fig1, curve_file)
    plt.close(fig1)

    buf2 = None
    if not mc_results.empty:
        fig2 = plt.figure(figsize=(10,5))
        for col in mc_results.columns:
            plt.hist(mc_results[col], bins=30, alpha=0.5, label=col, edgecolor="k")
        plt.axvline(mc_results.median().mean(), color="red", linestyle="--", label="Median")
        plt.title("Monte Carlo Final Equity Distribution (R)")
        plt.legend()
        plt.grid(True)
        buf2 = save_chart_as_image(fig2, mc_file)
        plt.close(fig2)

    # === Embed images ===
    if embed_charts:
        wb = load_workbook(report_file)
        if "Equity" in wb.sheetnames:
            ws = wb["Equity"]
            img1 = XLImage(buf1)
            img1.anchor = "H2"
            ws.add_image(img1)
        if buf2 and "MonteCarlo" in wb.sheetnames:
            ws = wb["MonteCarlo"]
            img2 = XLImage(buf2)
            img2.anchor = "H2"
            ws.add_image(img2)
        wb.save(report_file)
        print(f"Excel report saved with embedded charts → {report_file}")
    else:
        print(f"Excel report saved (no charts embedded) → {report_file}")

    print(f"Equity PNG saved → {curve_file}")
    if buf2:
        print(f"Monte Carlo PNG saved → {mc_file}")

    return trades, df_summary, df_mc_summary

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Secker Strategy Backtester")
    parser.add_argument("--mc", type=int, help="Monte Carlo runs (e.g. 1000)")
    parser.add_argument("--ruin", type=float, help="Global ruin threshold in R (e.g. -10)")
    parser.add_argument("--no-embed", action="store_true", help="Disable embedding charts in Excel")
    args = parser.parse_args()

    # Load or auto-create config
    cfg = ensure_config()

    # CLI overrides config
    mc_runs = args.mc if args.mc is not None else cfg.get("montecarlo_runs", 1000)
    ruin_thr = args.ruin if args.ruin is not None else cfg.get("ruin_threshold", -10)
    embed_charts = not args.no_embed if args.no_embed else cfg.get("embed_charts", True)

    df = load_trades()
    trades = build_trade_pairs(df)
    ruin_thresholds = cfg.get("ruin_thresholds", DEFAULT_CONFIG["ruin_thresholds"])
    trades, summary, mc_summary = analyze_trades(
        trades,
        montecarlo_runs=mc_runs,
        ruin_threshold=ruin_thr,
        ruin_thresholds=ruin_thresholds,
        embed_charts=embed_charts
    )

