import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
import argparse
import io
import json, yaml
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from scipy.stats import ks_2samp
from scipy.spatial.distance import jensenshannon

LEDGER_FILE = "trades.csv"
CONFIG_FILE = "config.yaml"

DEFAULT_CONFIG = {
    "montecarlo_runs": 1000,
    "ruin_threshold": -10,
    "embed_charts": True,
    "ruin_thresholds": {
        "180PC": -6,
        "T-Wave": -12,
        "Breakfast": -8
    },
    # Compounding defaults
    "start_balance": 10000.0,
    "risk_fraction": 0.01,
    "commission": 0.0,
    "slippage_bps": 0.0,
    # Monte Carlo compounded
    "mc_compounded_runs": 1000,
    "ruin_money_floor": 2000.0,
    "ruin_max_dd_pct": 60.0
}

def ensure_config():
    cfg_path = Path(CONFIG_FILE)
    if not cfg_path.exists():
        with open(cfg_path, "w") as f:
            yaml.safe_dump(DEFAULT_CONFIG, f)
        print(f"⚙️  Default config created → {CONFIG_FILE}")
        return DEFAULT_CONFIG
    else:
        with open(cfg_path, "r") as f:
            if cfg_path.suffix == ".json":
                return json.load(f)
            else:
                return yaml.safe_load(f)

def load_trades(file=LEDGER_FILE):
    return pd.read_csv(file, parse_dates=["timestamp"])

def build_trade_pairs(df: pd.DataFrame):
    trades, open_trades = [], {}
    for _, row in df.iterrows():
        strat = row["strategy"]
        if row["event"] == "entry":
            open_trades[strat] = {
                "timestamp": row["timestamp"],
                "strategy": strat,
                "dir": row["dir"],
                "entry": row["entry"],
                "sl": row["sl"],
                "tp": row["tp"],
            }
        elif row["event"] in ["sl", "tp", "trail"]:
            if strat in open_trades:
                tr = open_trades.pop(strat)
                tr["exit_event"] = row["event"]
                tr["exit_price"] = row["price"]
                trades.append(tr)
    return pd.DataFrame(trades)

def compute_drawdown(equity):
    peaks = np.maximum.accumulate(equity)
    dd = equity - peaks
    max_dd = dd.min() if len(dd) else 0
    durations, cur_dur = [], 0
    for d in dd:
        if d < 0: cur_dur += 1
        elif cur_dur > 0:
            durations.append(cur_dur)
            cur_dur = 0
    avg_dd_dur = np.mean(durations) if durations else 0
    return max_dd, avg_dd_dur

def performance_ratios(R_series):
    R = np.array(R_series.dropna())
    if len(R) == 0:
        return (None, None, None)
    gross_profit = R[R > 0].sum()
    gross_loss = -R[R < 0].sum()
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else np.inf
    sharpe = np.mean(R) / np.std(R) if np.std(R) > 0 else None
    downside = R[R < 0]
    sortino = np.mean(R) / np.std(downside) if len(downside) > 0 and np.std(downside) > 0 else None
    return (profit_factor, sharpe, sortino)

def compounding_path(R_series, start_balance=10000.0, risk_fraction=0.01,
                     commission=0.0, slippage_bps=0.0):
    R = np.asarray(pd.Series(R_series).fillna(0.0).values, dtype=float)
    slip_R = (slippage_bps / 10000.0) / max(risk_fraction, 1e-12)
    R_net = R - slip_R
    n = len(R_net)
    equity = np.empty(n+1, dtype=float)
    equity[0] = start_balance
    ret = np.empty(n, dtype=float)
    money_pnl = np.empty(n, dtype=float)
    for i in range(n):
        risk_amt = equity[i] * risk_fraction
        pnl_i = risk_amt * R_net[i] - commission
        money_pnl[i] = pnl_i
        equity[i+1] = equity[i] + pnl_i
        ret[i] = pnl_i / equity[i]
    eq_series = pd.Series(equity[1:], name="Equity")
    ret_series = pd.Series(ret, name="Return")
    pnl_series = pd.Series(money_pnl, name="MoneyPnL")
    peaks = eq_series.cummax()
    dd = (eq_series - peaks) / peaks
    max_dd_pct = dd.min() if len(dd) else 0.0
    trades_per_year = 252
    years = max(len(ret) / trades_per_year, 1e-12)
    cagr = (eq_series.iloc[-1] / start_balance) ** (1/years) - 1 if len(eq_series) else 0.0
    vol_ann = ret_series.std(ddof=0) * np.sqrt(trades_per_year) if len(ret_series) else 0.0
    mar = (cagr / abs(max_dd_pct)) if max_dd_pct < 0 else np.inf
    comp_df = pd.DataFrame({
        "MoneyPnL": pnl_series,
        "Return": ret_series,
        "Equity": eq_series,
        "DrawdownPct": dd
    })
    summary = {
        "FinalEquity": float(eq_series.iloc[-1]) if len(eq_series) else start_balance,
        "CAGR": float(cagr),
        "MaxDD%": float(max_dd_pct * 100.0),
        "MAR": float(mar),
        "AnnVol%": float(vol_ann * 100.0)
    }
    return comp_df, summary

def monte_carlo_compounded_equity(
    R_series,
    runs=1000,
    start_balance=10000.0,
    risk_fraction=0.01,
    commission=0.0,
    slippage_bps=0.0,
    ruin_money_floor=None,
    ruin_max_dd_pct=None
):
    R = np.asarray(pd.Series(R_series).dropna().values, dtype=float)
    n = len(R)
    if n == 0 or runs <= 0:
        return [], {"RuinProb%": 0.0, "MedianFinal": None, "P05": None, "P95": None}, 0
    slip_R = (slippage_bps / 10000.0) / max(risk_fraction, 1e-12)
    finals, ruin = [], 0
    for _ in range(runs):
        R_shuf = R[np.random.permutation(n)] - slip_R
        eq, peak, ruined = start_balance, start_balance, False
        for Ri in R_shuf:
            risk_amt = eq * risk_fraction
            pnl = risk_amt * Ri - commission
            eq += pnl
            if eq > peak: peak = eq
            if ruin_money_floor is not None and eq <= ruin_money_floor:
                ruined = True
            if ruin_max_dd_pct is not None and peak > 0:
                dd_pct = (eq - peak) / peak * 100.0
                if dd_pct <= -abs(ruin_max_dd_pct):
                    ruined = True
        finals.append(eq)
        if ruined: ruin += 1
    finals = np.array(finals, dtype=float)
    mc_summary = {
        "RuinProb%": float(ruin / runs * 100.0),
        "MedianFinal": float(np.median(finals)),
        "P05": float(np.percentile(finals, 5)),
        "P95": float(np.percentile(finals, 95))
    }
    return finals.tolist(), mc_summary, ruin

def js_divergence(a, b, bins=50):
    a_hist, bin_edges = np.histogram(a, bins=bins, density=True)
    b_hist, _ = np.histogram(b, bins=bin_edges, density=True)
    a_hist = a_hist + 1e-12
    b_hist = b_hist + 1e-12
    a_hist = a_hist / a_hist.sum()
    b_hist = b_hist / b_hist.sum()
    return jensenshannon(a_hist, b_hist)

def analyze_trades(trades: pd.DataFrame,
                   montecarlo_runs=1000,
                   ruin_threshold=-10,
                   ruin_thresholds=None,
                   embed_charts=True,
                   start_balance=10000.0,
                   risk_fraction=0.01,
                   commission=0.0,
                   slippage_bps=0.0,
                   mc_compounded_runs=1000,
                   ruin_money_floor=None,
                   ruin_max_dd_pct=None):
    trades = trades.copy()
    trades["risk"] = (trades["entry"] - trades["sl"]).abs()
    trades["pnl"] = trades.apply(
        lambda r: (r["exit_price"] - r["entry"]) if r["dir"] == "long"
        else (r["entry"] - r["exit_price"]), axis=1)
    trades["R"] = trades["pnl"] / trades["risk"]

    summary = []
    trades = trades.sort_values("timestamp")
    trades["portfolio_equity"] = trades["R"].cumsum()
    max_dd, avg_dd_dur = compute_drawdown(trades["portfolio_equity"].values)
    winrate = (trades["R"] > 0).mean()
    expectancy = trades["R"].mean()
    pf, sharpe, sortino = performance_ratios(trades["R"])
    summary.append({
        "Strategy": "Portfolio",
        "Trades": len(trades),
        "WinRate%": winrate*100,
        "ExpectancyR": expectancy,
        "ProfitFactor": pf,
        "Sharpe": sharpe,
        "Sortino": sortino,
        "MaxDD(R)": max_dd,
        "AvgDD_Dur": avg_dd_dur,
        "RuinProb%": None,
    })

    grouped = trades.groupby("strategy")
    for strat, g in grouped:
        g = g.sort_values("timestamp")
        g["equity"] = g["R"].cumsum()
        max_dd, avg_dd_dur = compute_drawdown(g["equity"].values)
        winrate = (g["R"] > 0).mean()
        expectancy = g["R"].mean()
        pf, sharpe, sortino = performance_ratios(g["R"])
        summary.append({
            "Strategy": strat,
            "Trades": len(g),
            "WinRate%": winrate*100,
            "ExpectancyR": expectancy,
            "ProfitFactor": pf,
            "Sharpe": sharpe,
            "Sortino": sortino,
            "MaxDD(R)": max_dd,
            "AvgDD_Dur": avg_dd_dur,
            "RuinProb%": None,
        })

    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    report_file = f"strategy_report_{ts}.xlsx"

    # Compounding
    comp_df, comp_sum = compounding_path(
        trades["R"], start_balance=start_balance,
        risk_fraction=risk_fraction, commission=commission,
        slippage_bps=slippage_bps
    )
    summary[0].update(comp_sum)

    # Monte Carlo Compounded
    mc_final_eq, mc_comp_summary, _ = monte_carlo_compounded_equity(
        trades["R"],
        runs=mc_compounded_runs,
        start_balance=start_balance,
        risk_fraction=risk_fraction,
        commission=commission,
        slippage_bps=slippage_bps,
        ruin_money_floor=ruin_money_floor,
        ruin_max_dd_pct=ruin_max_dd_pct
    )

    # For comparison: fake R-based MC distribution (cumsum of shuffled Rs)
    results_R = []
    for _ in range(montecarlo_runs):
        R_shuf = trades["R"].sample(frac=1, replace=False).cumsum().iloc[-1]
        results_R.append(R_shuf)

    # KS + JS metrics
    ks_stat, ks_pval = ks_2samp(results_R, mc_final_eq) if len(results_R) and len(mc_final_eq) else (None,None)
    jsd_val = js_divergence(results_R, mc_final_eq) if len(results_R) and len(mc_final_eq) else None

    # DataFrames
    df_summary = pd.DataFrame(summary)
    df_mc_comp_summary = pd.DataFrame([mc_comp_summary])
    df_mc_comp_dist = pd.DataFrame({"FinalEquity": mc_final_eq})
    df_mc_compare = pd.DataFrame([{
        "Mode": "Portfolio",
        "Runs": montecarlo_runs,
        "RuinThreshold(R)": ruin_threshold,
        "RuinProb%(R)": None,
        "RuinFloor($)": ruin_money_floor,
        "RuinMaxDD%": ruin_max_dd_pct,
        "RuinProb%(Money)": mc_comp_summary["RuinProb%"],
        "MedianFinal($)": mc_comp_summary["MedianFinal"],
        "P05($)": mc_comp_summary["P05"],
        "P95($)": mc_comp_summary["P95"],
        "KS_Distance": ks_stat,
        "KS_pValue": ks_pval,
        "JS_Divergence": jsd_val
    }])

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        trades.to_excel(writer, sheet_name="Trades", index=False)
        comp_df.to_excel(writer, sheet_name="Compounding", index=False)
        df_mc_comp_summary.to_excel(writer, sheet_name="MC_Compounding", index=False)
        df_mc_comp_dist.to_excel(writer, sheet_name="MCc_Distribution", index=False)
        df_mc_compare.to_excel(writer, sheet_name="MC_Compare", index=False)

    # Charts
    comp_curve_file = f"compounded_equity_{ts}.png"
    plt.figure(figsize=(12,6))
    plt.plot(comp_df.index, comp_df["Equity"], marker="o")
    plt.title("Compounded Equity (Money Terms)")
    plt.grid(True)
    plt.savefig(comp_curve_file, dpi=300, bbox_inches="tight")
    plt.close()

    mc_comp_file = f"mc_compounded_histogram_{ts}.png"
    plt.figure(figsize=(10,5))
    plt.hist(df_mc_comp_dist["FinalEquity"].values, bins=30, edgecolor="k")
    plt.title("Monte Carlo Final Equity Distribution (Compounded, $)")
    plt.grid(True)
    plt.savefig(mc_comp_file, dpi=300, bbox_inches="tight")
    plt.close()

    mc_dual_file = f"mc_dual_compare_{ts}.png"
    plt.figure(figsize=(10,6))
    plt.hist(results_R, bins=30, alpha=0.5, color="blue", edgecolor="k",
             label="R-based (normalized)", density=True)
    plt.hist(mc_final_eq, bins=30, alpha=0.5, color="green", edgecolor="k",
             label="Money-based (normalized)", density=True)
    plt.title(f"Monte Carlo Final Equity Distribution: R vs Money\nKS={ks_stat:.3f}, JS={jsd_val:.3f}")
    plt.xlabel("Final Equity")
    plt.ylabel("Density")
    plt.legend()
    plt.grid(True)
    plt.savefig(mc_dual_file, dpi=300, bbox_inches="tight")
    plt.close()

    if embed_charts:
        wb = load_workbook(report_file)
        if "Compounding" in wb.sheetnames:
            ws = wb["Compounding"]
            img = XLImage(comp_curve_file)
            img.anchor = "H2"
            ws.add_image(img)
        if "MC_Compounding" in wb.sheetnames:
            ws = wb["MC_Compounding"]
            img = XLImage(mc_comp_file)
            img.anchor = "H2"
            ws.add_image(img)
        if "MC_Compare" in wb.sheetnames:
            ws = wb["MC_Compare"]
            img = XLImage(mc_dual_file)
            img.anchor = "H20"
            ws.add_image(img)
        wb.save(report_file)

    print(f"Excel report saved → {report_file}")
    print(f"Compounded equity chart → {comp_curve_file}")
    print(f"MC compounded histogram → {mc_comp_file}")
    print(f"MC dual comparison chart → {mc_dual_file}")
    return trades, df_summary

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Secker Strategy Backtester")
    parser.add_argument("--mc", type=int, help="Monte Carlo runs (R-based)")
    parser.add_argument("--ruin", type=float, help="Global ruin threshold in R")
    parser.add_argument("--no-embed", action="store_true", help="Disable embedding charts in Excel")
    args = parser.parse_args()

    cfg = ensure_config()
    mc_runs = args.mc if args.mc is not None else cfg.get("montecarlo_runs", 1000)
    ruin_thr = args.ruin if args.ruin is not None else cfg.get("ruin_threshold", -10)
    embed_charts = not args.no_embed if args.no_embed else cfg.get("embed_charts", True)

    df = load_trades()
    trades = build_trade_pairs(df)
    trades, summary = analyze_trades(
        trades,
        montecarlo_runs=mc_runs,
        ruin_threshold=ruin_thr,
        ruin_thresholds=cfg.get("ruin_thresholds", DEFAULT_CONFIG["ruin_thresholds"]),
        embed_charts=embed_charts,
        start_balance=cfg.get("start_balance", 10000.0),
        risk_fraction=cfg.get("risk_fraction", 0.01),
        commission=cfg.get("commission", 0.0),
        slippage_bps=cfg.get("slippage_bps", 0.0),
        mc_compounded_runs=cfg.get("mc_compounded_runs", 1000),
        ruin_money_floor=cfg.get("ruin_money_floor", None),
        ruin_max_dd_pct=cfg.get("ruin_max_dd_pct", None)
    )